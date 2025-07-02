# This file contains the code to create the excel report for the monthly forecast
# Last update: 08.14.2023 by Inflation Forecasting Farotimi

# Importing necessary modules
import os
import subprocess
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime
from pathlib import Path
import sys
#statsmodels - for time series analysis
import statsmodels.stats.api as sms
from statsmodels.formula.api import ols
import plotly.graph_objects as go
import plotly.express as px
import plotly.io as pio

#openpyxl - for excel
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

#Python Image Library   
from PIL import Image as PILImage
from PIL import ImageOps

#import custom functions
from monthly_forecast_run import run_monthly_forecast
from models import arima, sarimax, varmax
from sklearn.metrics import mean_squared_error

#test residuals for heteroskedasticity 

TEST_NAMES = ['White', 'Breusch-Pagan', 'Goldfeld-Quandt', 'ARCH']
FORMULA = 'value ~ time'

class Heteroskedasticity:

    @staticmethod
    def het_tests(series: pd.Series, test: str) -> float:
        """
        Testing for heteroskedasticity

        :param series: Univariate time series as pd.Series
        :param test: String denoting the test. One of 'white','goldfeldquandt', or 'breuschpagan'

        :return: p-value as a float.

        If the p-value is high, we accept the null hypothesis that the data is homoskedastic
        """
        assert test in TEST_NAMES, 'Unknown test'

        series = series.reset_index(drop=True).reset_index()
        series.columns = ['time', 'value']
        series['time'] += 1

        olsr = ols(FORMULA, series).fit()

        if test == 'White':
            _, p_value, _, _ = sms.het_white(olsr.resid, olsr.model.exog)

        elif test == 'Goldfeld-Quandt':
            _, p_value, _ = sms.het_goldfeldquandt(olsr.resid, olsr.model.exog, alternative='two-sided')
        elif test == 'Breusch-Pagan':
            _, p_value, _, _ = sms.het_breuschpagan(olsr.resid, olsr.model.exog)
        elif test == 'ARCH':
            _, p_value, _, _ = sms.het_arch(olsr.resid)
            
    
        return p_value

    @classmethod
    def run_all_tests(cls, series: pd.Series):

        test_results = {k: cls.het_tests(series, k) for k in TEST_NAMES}

        return test_results

    

def create_aggregate_report(data):


    """ 
    Create an Excel file with the 12-month forecast, test set results and metrics, and model summary statistics for each country.
    It also includes a chart of the forecast and a chart of the test set results. The report is saved in the reports folder. 
    :param data: Object containing the exogenous variable daata in the following format: 
    {name_of_country: dataframe_with_data, name_of_country: dataframe_with_exogenous_variable_time_series, ...}
    :param countries: (list, optional): List of countries to include in the report. If empty, all countries in data will be included.
  
    Example:
        create_forecast_report(data, countries = ['CHL', 'MEX', 'ARG', 'BOL', 'BRA', 'PER'])
        data = {'CHL': chl_data, 'MEX': mex_data, 'ARG': arg_data, 'BOL': bol_data, 'BRA': bra_data, 'PER': per_data}
        where chl_data, mex_data, arg_data, bol_data, bra_data, and per_data are dataframes with the following columns: 'Inflation', and Date (index)

    """
    #get model name 
    multivar_data = {}

    for key in data.keys():
        if data[key].shape[1] > 1:
            multivar_data[key] = data[key]



    #get year, momnth, day for file name
    today = datetime.today().strftime('%Y-%m-%d')
    #GET TODAY'S MONTH NAME
    today_month = datetime.today().strftime('%B')
    today_year = datetime.today().strftime('%Y')

    #call run_monthly_forecast function to get forecast results for each country
    if data.isinstance(dict):
        arima_results = run_monthly_forecast(arima, data, data.keys())


        arimax_results = run_monthly_forecast(sarimax, multivar_data, multivar_data.keys())
        varmax_results = run_monthly_forecast(varmax, multivar_data, multivar_data.keys())
            
        
        aggregate_forecast = [arima_results, arimax_results, varmax_results]
    
    else:
        aggregate_forecast = list(data['Results Set'])


    #close excel if open
    os.system("taskkill /f /im EXCEL.EXE")

    #set color code variables 
    fill_good = "C6EFCE"
    fill_bad = "FFC7CE"
    fill_neutral = "FFEB9C"

    #set font color variables
    font_good = "006100"
    font_bad = "9C0006"
    font_neutral = "9C6500"

    #set border style
    rrange_border = Border(left=Side(style='thin'),     
                        right=Side(style='thin'), 
                        top=Side(style='thin'),  
                        bottom=Side(style='thin'))
    


    #create a new excel workbook
    wb = openpyxl.Workbook()
    #delete sheet 1
    wb.remove(wb['Sheet'])

    ''' 
    #create summary sheet
    summary = wb.create_sheet("Forecasts - All")
    #format fill background to white
    #add row and column 
    summary.insert_cols(1)
    summary.insert_rows(1)

    #set worksheet background fill to white
    for row in summary.iter_rows(max_row = 300, max_col = 15):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="FFFFFF")

    #set row width
    summary.column_dimensions['A'].width = 5
    summary.column_dimensions['B'].width = 10
    summary.column_dimensions['C'].width = 20
    summary.column_dimensions['D'].width = 20 
    summary.column_dimensions['E'].width = 35
    summary.column_dimensions['F'].width = 20
    summary.column_dimensions['G'].width = 20

    #sort forecasts by country
    #aggregate_forecast.sort(key=lambda x: x['country'])

    #create report for each country forecast in forecast_results
    ''' 
    country_start_row = 2

    # BEGIN: 5f5d8e9d6c7d
    forecast_dict = {}
    for model_results in aggregate_forecast:
        for country_results in model_results:
            country = country_results['country']
            if country not in forecast_dict:
                forecast_dict[country] = []
            forecast_dict[country].append(country_results)
    # END: 5f5d8e9d6c7d


    for country in forecast_dict:
        data  = forecast_dict[country][0]['data']

        naive_rmse = forecast_dict[country][0]['naive_rmse']
        ''' 
        #add country forecast table to summary sheet
        summary['B' + str(country_start_row)] = forecast['country'] + " - " + forecast['indicator'] + " - " + "Next 12 months forecast as of:  " + data.index[-1].strftime("%Y-%m-%d") 
        summary.merge_cells('B' + str(country_start_row) +  ':G' + str(country_start_row))

        
        summary['B' + str(2 + country_start_row)] = "Date"
        summary['C' + str(2 + country_start_row)] = "Inflation Change YoY %"
        summary['D' + str(2 + country_start_row)] = "Lower Bound - 95% Confidence Interval"
        summary['E' + str(2 + country_start_row)] = "Upper Bound - 95% Confidence Interval"

        

        #format titles
        summary['B' + str(country_start_row)].font = Font(color="366092", bold=True, size = 18)

        summary['B' + str(2 + country_start_row )].font = Font(bold = True)
        summary['C' + str(2 + country_start_row )].font = Font(bold = True)
        summary['D' + str(2 + country_start_row )].font = Font(bold = True)
        summary['E' + str(2 + country_start_row )].font = Font(bold = True)

        #format fill background to light blue
        summary['B' + str(2 + country_start_row)].fill = PatternFill("solid", fgColor="DCE6F1")
        summary['C' + str(2 + country_start_row)].fill = PatternFill("solid", fgColor="DCE6F1")
        summary['D' + str(2 + country_start_row)].fill = PatternFill("solid", fgColor="DCE6F1")
        summary['E' + str(2 + country_start_row)].fill = PatternFill("solid", fgColor="DCE6F1")

        #format fill background to light blue

         '''


        forecast_values = {}

        for model_results in forecast_dict[country]:
            forecast_values[model_results['model']] = model_results['forecast'].predicted_mean

        '''
        for i in range(0,len(forecast_mean)):
            summary['B'+str(3 + i + country_start_row)] = forecast_mean.index[i].strftime("%Y-%m-%d")
            summary['C'+str(3 + i + country_start_row)] = round(forecast_mean[i],2)
            #summary['D'+str(3 + i + country_start_row)] = round(np.exp(forecast['forecast'].conf_int().iloc[i,0]),2)
            #summary['E'+str(3 + i + country_start_row)] = round(np.exp(forecast['forecast'].conf_int().iloc[i,1]),2)
            summary['D'+str(3 + i + country_start_row)] = round(forecast['conf_int']['lower'].iloc[i],2)
            summary['E'+str(3 + i + country_start_row)] = round(forecast['conf_int']['upper'].iloc[i],2)
        
        
        #add borders 
        for row in summary.iter_rows(min_row= 2 + country_start_row, max_row=2 + country_start_row + len(forecast_mean), min_col=2, max_col=5):
                for cell in row:
                    cell.border = rrange_border


        country_start_row += 16

        '''

        #create a new worksheet for each country
        ws = wb.create_sheet(forecast_dict[country][0]['country'])
 
        #write forecast results to excel
        ws['A1'] = forecast_dict[country][0]['country'] + " - " + forecast_dict[country][0]['indicator'] + " - " + "Next 12 months forecast as of:  " + data.index[-1].strftime("%Y-%m-%d") 
        ws.merge_cells('A1:G1')
        ws['A3'] = "Models" 
        try:
            ws['B3'] = forecast_dict[country][0]['model'] +" - " + str(forecast_dict[country][0]['params'])
            ws['C3'] = forecast_dict[country][1]['model'] + " - " + str(forecast_dict[country][1]['params'])
            ws['D3'] = forecast_dict[country][2]['model'] + " - " + str(forecast_dict[country][2]['params'])
        except IndexError:
            pass
        try:
            ws['A4'] = "Training Data"
            ws['B4'] = forecast_dict[country][0]['data'].index[0].strftime("%Y-%m-%d") + " to " + forecast_dict[country][0]['data'].index[-12].strftime("%Y-%m-%d")
            ws['C4'] = forecast_dict[country][1]['data'].index[0].strftime("%Y-%m-%d") + " to " + forecast_dict[country][1]['data'].index[-12].strftime("%Y-%m-%d")
            ws['D4'] = forecast_dict[country][2]['data'].index[0].strftime("%Y-%m-%d") + " to " + forecast_dict[country][2]['data'].index[-12].strftime("%Y-%m-%d")
        except IndexError:
            pass

        try:
            #varmax
            ws['A5'] = "Test Data"
            ws['B5'] = forecast_dict[country][0]['data'].index[-12].strftime("%Y-%m-%d") + " to " + forecast_dict[country][0]['data'].index[-1].strftime("%Y-%m-%d")
            ws['C5'] = forecast_dict[country][1]['data'].index[-12].strftime("%Y-%m-%d") + " to " + forecast_dict[country][1]['data'].index[-1].strftime("%Y-%m-%d")
            ws['D5'] = forecast_dict[country][2]['data'].index[-12].strftime("%Y-%m-%d") + " to " + forecast_dict[country][2]['data'].index[-1].strftime("%Y-%m-%d")
        except IndexError:
            pass

        try:
            ws['A7'] = "Date"
            ws['B7'] = forecast_dict[country][0]['model']
            ws['C7'] = forecast_dict[country][1]['model']
            ws['D7'] = forecast_dict[country][2]['model']
        except IndexError:
            pass
        
        for i in range(len(forecast_values['ARIMA'])):
            ws['A'+str(8+i)] = forecast_values['ARIMA'].index[i].strftime("%Y-%m-%d")
            ws['B'+str(8+i)] = round(forecast_values.get('ARIMA', [0]*len(forecast_values['ARIMA']))[i],2)
            ws['C'+str(8+i)] = round(forecast_values.get('SARIMAX', [0]*len(forecast_values['ARIMA']))[i],2)
            ws['D'+str(8+i)] = round(forecast_values.get('VARMAX', [0]*len(forecast_values['ARIMA']))[i],2)
                
        #add forecast chart
        fc_chart = create_forecast_chart(forecast_dict[country])
        fc_chart_img = Image(fc_chart)

        ws.add_image(fc_chart_img,'B'+str(len(forecast_values['ARIMA'])+12))


        #add metrics
        ws['A54'] = "Test Metrics"
        #merge a54 and b54
        ws.merge_cells('A54:B54')
        ws['A55'] = "ARIMA" 
        try:
            ws['B55'] = str(forecast_dict[country][0]['test_metrics']['MAPE']) + "%"
        except IndexError:
            ws['B55'] = ""
        ws['A56'] = "SARIMAX"  
        try:
            ws['B56'] = str(forecast_dict[country][1]['test_metrics']['MAPE']) + "%"
        except IndexError:
            ws['B56'] = ""
        ws['A57'] = "VARMAX"
        try:
            ws['B57'] = str(forecast_dict[country][2]['test_metrics']['MAPE']) + "%"
        except IndexError:
            ws['B57'] = ""

        #add test predictions
        ws['A59'] = "Test Predictions"
        #merge a59 to e59
        #ws.merge_cells('A59:E59')
        ws['A60'] = "Date"
        ws['B60'] = "Actual"
        ws['C60'] = "Predicted (ARIMA)"
        ws['D60'] = "Predicted (ARIMAX)"
        ws['E60'] = "Predicted (VARMAX)"

        #add test predictions for test set
        for i in range(len(forecast_dict[country][0]['test_predictions'])):
            ws['A'+str(61+i)] = forecast_dict[country][0]['test_predictions'].index[i].strftime("%Y-%m-%d")
            ws['B'+str(61+i)] = round(forecast_dict[country][0]['test_actual'].iloc[i],2)
            ws['C'+str(61+i)] = round(forecast_dict[country][0]['test_predictions'].iloc[i],2)
            ws['D'+str(61+i)] = round(forecast_dict[country][1]['test_predictions'].iloc[i],2) if len(forecast_dict[country]) > 1 else ""
            ws['E'+str(61+i)] = round(forecast_dict[country][2]['test_predictions'].iloc[i],2) if len(forecast_dict[country]) > 2 else ""

        #add test chart
        test_chart = Image(create_test_chart(forecast_dict[country]))
        ws.add_image(test_chart,'B77')


       
        # Set the text in cells left-aligned
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column): 
            for cell in row: 
                cell.alignment = Alignment(horizontal='left')
    
        


        # set borders
        for row in ws.iter_rows(min_row=3, max_row=5, min_col=1, max_col=4):
                for cell in row:
                    cell.border = rrange_border

        for row in ws.iter_rows(min_row=7, max_row=19, min_col=1, max_col=4):
                for cell in row:
                        cell.border = rrange_border

        for row in ws.iter_rows(min_row=54, max_row=57, min_col=1, max_col=2):
                for cell in row:
                    cell.border = rrange_border

        for row in ws.iter_rows(min_row=59, max_row=72, min_col=1, max_col=5):
                for cell in row:
                    cell.border = rrange_border
        
        #set column width to fit content
        column_letters = tuple(get_column_letter(col_number + 1) for col_number in range(ws.max_column))
        for column_letter in column_letters:
            ws.column_dimensions[column_letter].width = 35

        #insert column 
        ws.insert_cols(1)
        ws.insert_rows(1)

        #set row width
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20

        #set worksheet background fill to white
        for row in ws.iter_rows(max_row = 300, max_col = 15):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FFFFFF")


        #set font and color for cell B2 to blue and bold
        ws['B2'].font = Font(color="366092", bold=True, size = 22)

        #set font and color for cells B4:B6 to bold
        for row in ws.iter_rows(min_row=4, max_row=6, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="DCE6F1")


        #set font and color for cells B8:E8 to bold
        for row in ws.iter_rows(min_row=8, max_row=8, min_col=2, max_col=5):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="DCE6F1")

        ''' 
        #set font and color for cells B49 to bold
        for row in ws.iter_rows(min_row=55, max_row=55, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(bold=True)

        #merge and center cells B4:B6
        ws.merge_cells('B55:C55')
        ws['B55'].alignment = Alignment(horizontal='center')
        ws['B55'].fill = PatternFill("solid", fgColor="DCE6F1")


        #merge and center cells B54:D55
        ws.merge_cells('B61:F61')
        ws['B61'].alignment = Alignment(horizontal='center')
        ws['B61'].fill = PatternFill("solid", fgColor="DCE6F1")
        ws['B61'].font = Font(bold=True)

        #merge and center cells B59:D59
        for row in ws.iter_rows(min_row=60, max_row=60, min_col=2, max_col=6):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="DCE6F1")

        

        #write model summary to excel
        summary_table_0 = forecast['model_summary'].tables[0].data
        summary_table_1 = forecast['model_summary'].tables[1].data
        summary_table_2 = forecast['model_summary'].tables[2].data

        ws.merge_cells('B102:C102')
        ws['B102'] = "Appendix A: Model Summary"
        ws['B102'].font = Font(bold=True)
        ws['B102'].fill = PatternFill("solid", fgColor="DCE6F1")

        #fill in model summary
        #general info
        ws['B103'] = "Model"
        ws['C103'] = summary_table_0[1][1]

        ws['B104'] = "Sample"
        ws['C104'] = summary_table_0[4][1] + summary_table_0[5][1]

        ws['B105'] = "No. of Observations"
        ws['C105'] = str(summary_table_0[0][3])

        #Terms and P_values
        ws['B106'] = "Terms"
        ws['C106'] = "P-value (Statistical significance)"

        # format header
        ws['B106'].font = Font(bold=True)
        ws['C106'].font = Font(bold=True)

        ws['B106'].fill = PatternFill("solid", fgColor="DCE6F1")
        ws['C106'].fill = PatternFill("solid", fgColor="DCE6F1")

        #format terms
        for i in range(1,len(summary_table_1)):
            ws['B' + str(106 + i)] = summary_table_1[i][0]
            ws['C' + str(106 + i)] = summary_table_1[i][4]

            if(float(summary_table_1[i][4]) < 0.05):
                #format good
                ws['C' + str(106 + i)].fill = PatternFill("solid", fgColor=fill_good)
                ws['C' + str(106 + i)].font = Font(color=font_good)
            else:
                #format bad
                ws['C' + str(106 + i)].fill = PatternFill("solid", fgColor=fill_bad)
                ws['C' + str(106 + i)].font = Font(color=font_bad)
           
        #add model assumptions header
        ws['B' + str(106 + len(summary_table_1))] = "Assumptions"
        ws['C' + str(106 + len(summary_table_1))] = "P-value"
        ws['D' + str(106 + len(summary_table_1))] = "Interpretation"

        # format header
        ws['B' + str(106 + len(summary_table_1))].font = Font(bold=True)
        ws['C' + str(106 + len(summary_table_1))].font = Font(bold=True)
        ws['D' + str(106 + len(summary_table_1))].font = Font(bold=True)

        ws['B' + str(106 + len(summary_table_1))].fill = PatternFill("solid", fgColor="DCE6F1")
        ws['C' + str(106 + len(summary_table_1))].fill = PatternFill("solid", fgColor="DCE6F1")
        ws['D' + str(106 + len(summary_table_1))].fill = PatternFill("solid", fgColor="DCE6F1")
        
        # model assumptions
        ws['B' + str(107 + len(summary_table_1))] = "Ljung-Box (L1)"
        ws['C' + str(107 + len(summary_table_1))] = summary_table_2[1][1]

        ws['B' + str(108 + len(summary_table_1))] = "Heteroskedasticity"
        ws['C' + str(108 + len(summary_table_1))] = summary_table_2[3][1]

        ws['B' + str(109 + len(summary_table_1))] = "Jarque-Bera"
        ws['C' + str(109 + len(summary_table_1))] = summary_table_2[1][3]

        ws['B' + str(110 + len(summary_table_1))] = "Skew"
        ws['C' + str(110 + len(summary_table_1))] = summary_table_2[2][3]

        ws['B' + str(111 + len(summary_table_1))] = "Kurtosis"
        ws['C' + str(111 + len(summary_table_1))] = summary_table_2[3][3]


        #interpret results
        if (float(summary_table_2[1][1]) > 0.05):
            ws['D' + str(107 + len(summary_table_1))] = "Residuals are independent."
            #format good
            ws['C' + str(107 + len(summary_table_1))].fill = PatternFill("solid", fgColor=fill_good)
            ws['C' + str(107 + len(summary_table_1))].font = Font(color=font_good)
        else:
            ws['D' + str(107 + len(summary_table_1))] = "Residuals are not independent."
            #format bad
            ws['C' + str(107 + len(summary_table_1))].fill = PatternFill("solid", fgColor=fill_bad)
            ws['C' + str(107 + len(summary_table_1))].font = Font(color=font_bad)


        if (float(summary_table_2[3][1]) > 0.05):
            ws['D' + str(108 + len(summary_table_1))] = "Residuals have the same variance."
            #format good
            ws['C' + str(108 + len(summary_table_1))].fill = PatternFill("solid", fgColor=fill_good)
            ws['C' + str(108 + len(summary_table_1))].font = Font(color=font_good)
        else:
            ws['D' + str(108 + len(summary_table_1))] = "Residuals have different variances."
            #format bad
            ws['C' + str(108 + len(summary_table_1))].fill = PatternFill("solid", fgColor=fill_bad)
            ws['C' + str(108 + len(summary_table_1))].font = Font(color=font_bad)
            
        if (float(summary_table_2[1][3]) > 0.05):
            ws['D' + str(109 + len(summary_table_1))] = "Residuals are normally distributed."
            #format good
            ws['C' + str(109 + len(summary_table_1))].fill = PatternFill("solid", fgColor=fill_good)
            ws['C' + str(109 + len(summary_table_1))].font = Font(color=font_good)
        else:
            ws['D' + str(109 + len(summary_table_1))] = "Residuals are not normally distributed."
            #format bad
            ws['C' + str(109 + len(summary_table_1))].fill = PatternFill("solid", fgColor=fill_bad)
            ws['C' + str(109 + len(summary_table_1))].font = Font(color=font_bad)

        if (float(summary_table_2[2][3]) > 0):  
            ws['D' + str(110 + len(summary_table_1))] = "Positive skew - model underestimates the mean."
            #format neutral
            ws['C' + str(110 + len(summary_table_1))].fill = PatternFill("solid", fgColor=fill_neutral)
            ws['C' + str(110 + len(summary_table_1))].font = Font(color=font_neutral)
        elif (float(summary_table_2[2][3]) < 0):                                                                                                       
            ws['D' + str(110 + len(summary_table_1))] = "Negative skew - Model overestimates the mean."
          #format neutral
            ws['C' + str(110 + len(summary_table_1))].fill = PatternFill("solid", fgColor=fill_neutral)
            ws['C' + str(110 + len(summary_table_1))].font = Font(color=font_neutral)

        elif (np.abs(float(summary_table_2[2][3])) >= 2):
            ws['C' + str(110 + len(summary_table_1))] = "Significant skew - over/under-estimation."
            #format bad
            ws['C' + str(110 + len(summary_table_1))].fill = PatternFill("solid", fgColor=fill_bad)
            ws['C' + str(110 + len(summary_table_1))].font = Font(color=font_bad)

        if (np.abs(float(summary_table_2[3][3])) < 3):
            ws['D' + str(111 + len(summary_table_1))] = "Residuals do not contain more extreme outliers than in a normal distribution."
            #format good
            ws['C' + str(111 + len(summary_table_1))].fill = PatternFill("solid", fgColor=fill_good)
            ws['C' + str(111 + len(summary_table_1))].font = Font(color=font_good)
        else:
            ws['D' + str(111 + len(summary_table_1))] = "Residuals contain more extreme outliers than in a normal distribution."
            #format bad
            ws['C' + str(111 + len(summary_table_1))].fill = PatternFill("solid", fgColor=fill_bad)
            ws['C' + str(111 + len(summary_table_1))].font = Font(color=font_bad)

            #borders
        for row in ws.iter_rows(min_row=102, max_row=111+len(summary_table_1), min_col=2, max_col=4):
                for cell in row:
                    cell.border = rrange_border


        #add plot of data with slice used highlighted 

        #add test chart
        data_chart = Image(forecast['orig_data_chart'])

        ws.add_image(data_chart,'B' + str(113 + len(summary_table_1)))

        #add residuals summary statistics
        ws['B' + str(140 + len(summary_table_1) + 1)] = "Appendix C: Residuals Analysis"
        ws['B' + str(140 + len(summary_table_1) + 1)].font = Font(bold=True)
        ws['B' + str(140 + len(summary_table_1) + 1)].fill = PatternFill("solid", fgColor="DCE6F1")

        ws['B' + str(140 + len(summary_table_1) + 2)] = "Mean"
        ws['C' + str(140 + len(summary_table_1) + 2)] = str(round(forecast['residuals'].mean(),2))
        ws['B' + str(140 + len(summary_table_1) + 3)] = "Variance"
        ws['C' + str(140 + len(summary_table_1) + 3)] = str(round(forecast['residuals'].std()**2,2))


        #CALCULATE white noise
        from statsmodels.stats.diagnostic import acorr_ljungbox
        ljung_box = acorr_ljungbox(forecast['residuals'], lags=12)
        p_values = ljung_box['lb_pvalue']
        if (p_values < 0.05).any():
            ws['B' + str(140 + len(summary_table_1) + 4)] = "White Noise?"
            ws['C' + str(140 + len(summary_table_1) + 4)] = "No" + " - Correlation at Lags " + str(np.where(p_values < 0.05)[0] + 1)
        else:
            ws['B' + str(140 + len(summary_table_1) + 4)] = "White Noise?"
            ws['C' + str(140 + len(summary_table_1) + 4)] = "Yes"


        #calculate heteroskedasticity of residuals
        het = Heteroskedasticity.run_all_tests(forecast['residuals'])


        #add heteroskedasticity results

        if (het['ARCH'] < 0.05):
            ws['B' + str(140 + len(summary_table_1) + 5)] = "Heteroskedastic? - ARCH"
            ws['C' + str(140 + len(summary_table_1) + 5)] = "Yes"
        else:
            ws['B' + str(140 + len(summary_table_1) + 5)] = "Heteroskedastic? - ARCH"  
            ws['C' + str(140 + len(summary_table_1) + 5)] = "No"



        
        #calculate normality and add results
        from scipy.stats import shapiro
        shapiro_test = shapiro(forecast['residuals'])
        if (shapiro_test[1] < 0.05):
            ws['B' + str(140 + len(summary_table_1) + 6)] = "Normal?"
            ws['C' + str(140 + len(summary_table_1) + 6)] = "No"
        else:
            ws['B' + str(140 + len(summary_table_1) + 6)] = "Normal?"
            ws['C' + str(140 + len(summary_table_1) + 6)] = "Yes"

        #add border
        ws.merge_cells('B' + str(140 + len(summary_table_1) + 1) + ':C' + str(140 + len(summary_table_1) + 1))
        for row in ws.iter_rows(min_row=140 + len(summary_table_1) + 1, max_row=140 + len(summary_table_1) + 6, min_col=2, max_col=3):
                for cell in row:
                    cell.border = rrange_border

        #color code white noise
        if (p_values < 0.05).any():
            ws['C' + str(140 + len(summary_table_1) + 4)].fill = PatternFill("solid", fgColor=fill_bad)
            ws['C' + str(140 + len(summary_table_1) + 4)].font = Font(color=font_bad)
        else:
            ws['C' + str(140 + len(summary_table_1) + 4)].fill = PatternFill("solid", fgColor=fill_good)
            ws['C' + str(140 + len(summary_table_1) + 4)].font = Font(color=font_good)

        #color code heteroskedasticity results
       
        if (het['ARCH'] < 0.05):
            ws['C' + str(140 + len(summary_table_1) + 5)].fill = PatternFill("solid", fgColor=fill_bad)
            ws['C' + str(140 + len(summary_table_1) + 5)].font = Font(color=font_bad)
        else:
            ws['C' + str(140 + len(summary_table_1) + 5)].fill = PatternFill("solid", fgColor=fill_good)
            ws['C' + str(140 + len(summary_table_1) + 5)].font = Font(color=font_good)

        #color code normality results
        if (shapiro_test[1] < 0.05):
            ws['C' + str(140 + len(summary_table_1) + 6)].fill = PatternFill("solid", fgColor=fill_bad)
            ws['C' + str(140 + len(summary_table_1) + 6)].font = Font(color=font_bad)
        else:
            ws['C' + str(140 + len(summary_table_1) + 6)].fill = PatternFill("solid", fgColor=fill_good)
            ws['C' + str(140 + len(summary_table_1) + 6)].font = Font(color=font_good)

        
        #add plot diagnostics 
        diagnostics = forecast['plot_diagnostics']

        #add title to diagnostics
        diagnostics.suptitle('Appendix D: Plot Diagnostics')
        diagnostics.tight_layout()
        diagnostics.set_size_inches(12, 6)

        diagnostics.savefig(('diagnostics_' + forecast['country'] + '.png'))

        #add border to image
        img = PILImage.open(('diagnostics_' + forecast['country'] + '.png'))
        bordered_img = ImageOps.expand(img, border=1, fill='black')
        bordered_img.save(('diagnostics_' + forecast['country'] + '.png'))

        #add plot diagnostics chart image to excel
        diagnostics = Image(('diagnostics_' + forecast['country'] + '.png'))
        ws.add_image(diagnostics, 'B' + str(151 + len(summary_table_1)))

        #plot acf of squared residuals - this is to check for ARCH
        from statsmodels.graphics.tsaplots import plot_acf
        acf_plot = plot_acf(forecast['residuals']**2, lags=12, alpha=0.05)
        acf_plot.figure.set_size_inches(12, 6)
        plt.suptitle('Appendix E: Autocorrelation Function of Squared Residuals')
        
        acf_plot.tight_layout()
        acf_plot.savefig(('acf_' + forecast['country'] + '.png'))

        #add border to image
        img2 = PILImage.open(('acf_' + forecast['country'] + '.png'))
        bordered_img2 = ImageOps.expand(img2, border=1, fill='black')
        bordered_img2.save(('acf_' + forecast['country'] + '.png'))

        #add acf chart image to excel
        acf = Image(('acf_' + forecast['country'] + '.png'))
        ws.add_image(acf, 'B' + str(184 + len(summary_table_1)))

        '''

       
        if len(forecast_dict[country]) > 0:
            if (float(forecast_dict[country][0]['test_metrics']['MAPE']) <= 20):
                ws['C56'].fill = PatternFill("solid",fgColor=fill_good) #GOOD
                ws['C56'].font = Font(color=font_good)
            elif(float(forecast_dict[country][0]['test_metrics']['RMSE']) > 20):
                ws['C56'].fill = PatternFill("solid",fgColor=fill_neutral) #NEUTRAL
                ws['C56'].font = Font(color=font_neutral)
            elif(float(forecast_dict[country][0]['test_metrics']['RMSE']) > 50):
                ws['C56'].fill = PatternFill("solid",fgColor=fill_bad) #BAD
                ws['C56'].font = Font(color=font_bad)

        if len(forecast_dict[country]) > 1:
            if (float(forecast_dict[country][1]['test_metrics']['MAPE']) <= 20):
                ws['C57'].fill = PatternFill("solid",fgColor=fill_good) #GOOD
                ws['C57'].font = Font(color=font_good)
            elif(float(forecast_dict[country][1]['test_metrics']['RMSE']) > 20):
                ws['C57'].fill = PatternFill("solid",fgColor=fill_neutral) #NEUTRAL
                ws['C57'].font = Font(color=font_neutral)
            elif(float(forecast_dict[country][1]['test_metrics']['RMSE']) > 50):
                ws['C57'].fill = PatternFill("solid",fgColor=fill_bad) #BAD
                ws['C57'].font = Font(color=font_bad)

    
        if len(forecast_dict[country]) > 2:
            if (float(forecast_dict[country][2]['test_metrics']['MAPE']) <= 20):
                ws['C58'].fill = PatternFill("solid",fgColor=fill_good) #GOOD
                ws['C58'].font = Font(color=font_good)
            elif(float(forecast_dict[country][2]['test_metrics']['RMSE']) > 20):
                ws['C58'].fill = PatternFill("solid",fgColor=fill_neutral) #NEUTRAL
                ws['C58'].font = Font(color=font_neutral)
            elif(float(forecast_dict[country][2]['test_metrics']['RMSE']) > 50):
                ws['C58'].fill = PatternFill("solid",fgColor=fill_bad) #BAD
                ws['C58'].font = Font(color=font_bad)
        
        

        #insert new row after row 4 and add exogenous variables 
        
           #add exogenous variables in C3 
        ws.insert_rows(7)
        for model in forecast_dict[country]:
            if(model['predictors'] is not None):
                exogenous_variables = model['predictors'].tolist()
                exogenous_variables = ', '.join(exogenous_variables)
                ws['B7'] = "Exogenous Variables"
                ws['C7'] = exogenous_variables


                #merge c7, c8 and c9 
                ws.merge_cells('C7:E7')

                #make border from c7 to e7
                for row in ws.iter_rows(min_row=7, max_row=7, min_col=2, max_col=5):
                    for cell in row:
                        cell.border = rrange_border

                ws['B7'].fill = PatternFill("solid", fgColor="DCE6F1")
                ws['B7'].font = Font(bold=True)
        
        #color code tab color based on MASE
        if all(float(forecast_dict[country][i]['test_metrics']['MAPE']) > 20 for i in range(len(forecast_dict[country]))):
            ws.sheet_properties.tabColor = fill_bad #BAD
        else:
            ws.sheet_properties.tabColor = fill_good #GOOD  


        #create chart by calling create_forecast_chart
        create_forecast_chart(forecast_dict[country])

    
    #format row 56
    ws.merge_cells('B56:C56')
    ws['B56'].alignment = Alignment(horizontal='center')
    ws['B56'].fill = PatternFill("solid", fgColor="DCE6F1")
    ws['B56'].font = Font(bold=True)

    #format row 61
    ws.merge_cells('B61:F61')
    ws['B61'].alignment = Alignment(horizontal='center')
    ws['B61'].fill = PatternFill("solid", fgColor="DCE6F1")
    ws['B61'].font = Font(bold=True)

    #format row 62
    #make b62 to f62 bold and fill with color
    for row in ws.iter_rows(min_row=62, max_row=62, min_col=2, max_col=6):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DCE6F1")




    #save the workbook
    FILE_PATH = r'reports' + f"\{today_year}\{today_month}\{today}\\"
    SAVE_PATH = f"{forecast_dict[country][0]['indicator']}_forecast_report_{datetime.today().strftime('%Y-%m-%d_%H_%M')}.xlsx"

    #create directory if it doesn't exist
    if not os.path.exists(FILE_PATH):
        os.makedirs(FILE_PATH)
    
    #save file
    wb.save(FILE_PATH + SAVE_PATH)

    
    
    #open saved file
    subprocess.Popen([FILE_PATH + SAVE_PATH], shell=True)

def create_forecast_chart(country_fc_dict):
    today_year = datetime.today().strftime('%Y')
    today_month = datetime.today().strftime('%B')
    today = datetime.today().strftime('%Y-%m-%d')
    country = country_fc_dict[0]['country']
    indicator = country_fc_dict[0]['indicator']
    target = country_fc_dict[0]['target']
      

    y = country_fc_dict[0]['data'][target]
    past_Y = y[-12:]

    forecast_values = {}
    #for each model, get the forecast and add to all_Y
    for i in range(0,len(country_fc_dict)):
        model_name = country_fc_dict[i]['model']
        forecast_values[model_name] = country_fc_dict[i]['forecast'].predicted_mean


    #create chart
    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x = past_Y.index, y = past_Y, name = 'Actual', line=dict(color='blue')))

    

    for model in forecast_values:
        fig3.add_trace(go.Scatter(x = forecast_values[model].index, y = forecast_values[model], name = model))

    fig3.update_layout(title = country_fc_dict[0]['country'] + " - " + country_fc_dict[0]['indicator'] + " - " + " 12 months forecast ", xaxis_title = 'Date', yaxis_title = country_fc_dict[0]['target'], height = 600, width = 1400)
    
    #add line for current date
    fig3.add_shape(type="line", yref='paper', x0=y.index[-1], y0=0, x1=y.index[-1], y1=1, line=dict(color="Red",width=1, dash="dot")) 
    
    
    

     
     



    #get metrics for test set evaluation
    '''
    rmse = country_fc_dict['test_metrics']['RMSE']
    aic = country_fc_dict['test_metrics']['AIC']
    mape = country_fc_dict['test_metrics']['MAPE']
    naive_rmse = country_fc_dict['test_metrics']['Naive RMSE']

    #add metrics annotation to chart with metrics
    fig3.add_annotation(xref='paper',
                        yref='paper',
                        yanchor="bottom",
                        borderpad=1,
                        x = 0.5,
                        y = 0.8,
                        font=dict(size=10),
                        bgcolor="white",
                        bordercolor='black',
                        borderwidth=1,
                        text="Model Evaluation Metrics - Training RMSE " + str("x") + " | RMSE: " + str(round(rmse,2)) + " | Naive RMSE: " + str(round(naive_rmse, 2)) + " | MAPE: " + str(round(mape,2)) + "%",
                        showarrow=False,
                        align="center")
    '''

    fig3.show()

    #save image
    CHART_PATH = r"reports/" + today_year + "/" + today_month + "/" + today + "/" + country + "/" + country +" _"+ indicator + '_forecastchart.png'
    parent_dir = Path(CHART_PATH).parent.absolute()
    
    import os

    # specify the path of the directory you want to create
    directory_path = "path/to/directory"

    # create directory if it doesn't exist
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)
    pio.write_image(fig3, file=CHART_PATH)

    #add border to image
    img = PILImage.open(CHART_PATH)
    bordered_img = ImageOps.expand(img, border=1, fill='black')
    bordered_img.save(CHART_PATH)

    return CHART_PATH

def create_test_chart(country_fc_dict):
    today_year = datetime.today().strftime('%Y')
    today_month = datetime.today().strftime('%B')
    today = datetime.today().strftime('%Y-%m-%d')
    country = country_fc_dict[0]['country']
    indicator = country_fc_dict[0]['indicator']
    target = country_fc_dict[0]['target']
      

    y = country_fc_dict[0]['data'][target]
    past_Y = y[-12:]

    test_predictions = {}
    #for each model, get the forecast and add to all_Y
    for i in range(0,len(country_fc_dict)):
        model_name = country_fc_dict[i]['model']
        test_predictions[model_name] = country_fc_dict[i]['test_predictions']
        test_actual = country_fc_dict[i]['test_actual']

    
    
    #create chart
    fig3 = go.Figure()

    #add actual
    fig3.add_trace(go.Scatter(x = test_actual.index, y = test_actual, name = 'Actual', line=dict(color='blue')))

    #add naive
    fig3.add_trace(go.Scatter(x = test_actual.index, y = test_actual.shift(1), name = 'Naive', line=dict(color='gold')))


    for model in test_predictions:
        fig3.add_trace(go.Scatter(x = test_predictions[model].index, y = test_predictions[model], name = model))

    
    fig3.update_layout(title = country_fc_dict[0]['country'] + " - " + country_fc_dict[0]['indicator'] + " - " + " Test Set Predictions", xaxis_title = 'Date', yaxis_title = country_fc_dict[0]['target'], height = 600, width = 1400)
    
    metrics = {}
    metrics['Naive'] = {
        'MAPE': country_fc_dict[0]['test_metrics']['Naive MAPE']
    }
    for model in country_fc_dict:
        metrics[model['model']] = {
            'RMSE': model['test_metrics']['RMSE'],
            'AIC': model['test_metrics']['AIC'],
            'MAPE': model['test_metrics']['MAPE'],
        }

    #add metrics annotation to chart with metrics
    metrics_summary = ""
    for i, model in enumerate(metrics):
        metrics_summary += f"{model}: MAPE={metrics[model]['MAPE']}%" + "<br>"

    fig3.add_annotation(x=0.5, y=0.95, xanchor='left', yanchor='top', xref='paper', yref='paper', text=metrics_summary, showarrow=False, font=dict(size=10), bordercolor='black', borderwidth=1, borderpad=1, bgcolor='white', align='left')
            #text=metrics_summary, showarrow=False, font=dict(size=10))

    #add naive mape to chart
    fig3.show()

    #save image
    CHART_PATH = r"reports/" + today_year + "/" + today_month + "/" + today + "/" + country + "/" + country +" _"+ indicator + '_testchart.png'
    parent_dir = Path(CHART_PATH).parent.absolute()
    
    import os

    # specify the path of the directory you want to create
    directory_path = "path/to/directory"

    # create directory if it doesn't exist
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)
    pio.write_image(fig3, file=CHART_PATH)

    #add border to image
    img = PILImage.open(CHART_PATH)
    bordered_img = ImageOps.expand(img, border=1, fill='black')
    bordered_img.save(CHART_PATH)

    return CHART_PATH
    


   