# This file contains the code to create the excel report for the monthly forecast
# Last update: 08.14.2023 by Inflation Forecasting Farotimi

# Importing necessary modules
from ast import If
import os
import subprocess
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime
from scipy.stats import kurtosis, skew

#statsmodels - for time series analysis
import statsmodels.stats.api as sms
from statsmodels.formula.api import ols

#openpyxl - for excel
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

#Python Image Library   
from PIL import Image as PILImage
from PIL import ImageOps

#import custom functions
from monthly_forecast_run import run_monthly_forecast
from models import arima, sarimax
from sklearn.metrics import mean_squared_error

#test residuals for heteroskedasticity 

TEST_NAMES = ['White', 'Breusch-Pagan', 'Goldfeld-Quandt', 'ARCH']
FORMULA = 'value ~ time'

class Heteroskedasticity:

    @staticmethod
    def het_tests(series: pd.Series, test: str) -> float:
        """
        Testing for heteroskedasticity

        :param series: Univariate time series as pd.Series
        :param test: String denoting the test. One of 'white','goldfeldquandt', or 'breuschpagan'

        :return: p-value as a float.

        If the p-value is high, we accept the null hypothesis that the data is homoskedastic
        """
        assert test in TEST_NAMES, 'Unknown test'

        series = series.reset_index(drop=True).reset_index()
        series.columns = ['time', 'value']
        series['time'] += 1

        olsr = ols(FORMULA, series).fit()

        if test == 'White':
            _, p_value, _, _ = sms.het_white(olsr.resid, olsr.model.exog)

        elif test == 'Goldfeld-Quandt':
            _, p_value, _ = sms.het_goldfeldquandt(olsr.resid, olsr.model.exog, alternative='two-sided')
        elif test == 'Breusch-Pagan':
            _, p_value, _, _ = sms.het_breuschpagan(olsr.resid, olsr.model.exog)
        elif test == 'ARCH':
            _, p_value, _, _ = sms.het_arch(olsr.resid)
            
    
        return p_value

    @classmethod
    def run_all_tests(cls, series: pd.Series):

        test_results = {k: cls.het_tests(series, k) for k in TEST_NAMES}

        return test_results
    

def create_forecast_report(data, countries = [], model = None, log_transform = False, walkforward = False):


    """ 
    Create an Excel file with the 12-month forecast, test set results and metrics, and model summary statistics for each country.
    It also includes a chart of the forecast and a chart of the test set results. The report is saved in the reports folder. 
    :param data: Object containing the exogenous variable daata in the following format: 
    {name_of_country: dataframe_with_data, name_of_country: dataframe_with_exogenous_variable_time_series, ...}
    :param countries: (list, optional): List of countries to include in the report. If empty, all countries in data will be included.
  
    Example:
        create_forecast_report(data, countries = ['CHL', 'MEX', 'ARG', 'BOL', 'BRA', 'PER']).,
        data = {'CHL': chl_data, 'MEX': mex_data, 'ARG': arg_data, 'BOL': bol_data, 'BRA': bra_data, 'PER': per_data}
        where chl_data, mex_data, arg_data, bol_data, bra_data, and per_data are dataframes with the following columns: 'Inflation', and Date (index)

    """
    #get model name 


    #forecast chart list 
    forecast_chart_list = []


    #get year, momnth, day for file name
    today = datetime.today().strftime('%Y-%m-%d')
    today_month = datetime.today().strftime('%B')
    today_year = datetime.today().strftime('%Y')

    #call run_monthly_forecast function to get forecast results for each country
    if isinstance(data,list):
        forecast_results = data
    else:
        forecast_results = run_monthly_forecast(model, data, countries, log_transform, walkforward)


    #close excel if open
    os.system("taskkill /f /im EXCEL.EXE")

    #set color code variables 
    fill_good = "C6EFCE"
    fill_bad = "FFC7CE"
    fill_neutral = "FFEB9C"

    #set font color variables
    font_good = "006100"
    font_bad = "9C0006"
    font_neutral = "9C6500"

    #set border style
    rrange_border = Border(left=Side(style='thin'),     
                        right=Side(style='thin'), 
                        top=Side(style='thin'),  
                        bottom=Side(style='thin'))
    


    #create a new excel workbook
    wb = openpyxl.Workbook()
    #delete sheet 1
    wb.remove(wb['Sheet'])

    #create summary sheet
    summary = wb.create_sheet("Forecasts - All")
    #format fill background to white
    #add row and column 
    summary.insert_cols(1)
    summary.insert_rows(1)

    #set worksheet background fill to white
    for row in summary.iter_rows(max_row = 300, max_col = 15):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="FFFFFF")

    #set row width
    summary.column_dimensions['A'].width = 5
    summary.column_dimensions['B'].width = 10


    #sort forecasts by country
    forecast_results.sort(key=lambda x: x['country'])

    #create report for each country forecast in forecast_results
    country_start_row = 2


      #add country forecast table to summary sheet
    summary['B2'] = "PROYECCIONES DE INFLACIÓN (DE AQUÍ A 12 MESES) - LATINOAMERICA Y EL CARIBE"
    #format title
    summary['B2'].font = Font(color="366092", bold=True, size = 22)
    summary.merge_cells('B2:Z2')
    summary['B2'].alignment = Alignment(horizontal='center')
    summary['B' + str(2 + country_start_row)] = "Fecha"
    #format bold and fill background to light blue
    summary['B' + str(2 + country_start_row)].fill = PatternFill("solid", fgColor="DCE6F1")
    summary['B' + str(2 + country_start_row)].font = Font(bold=True)
    

    

    all_forecast_data = pd.DataFrame()

    for forecast in forecast_results: 
        #merge data into all_forecast_data
        all_forecast_data = pd.concat([all_forecast_data, forecast['forecast']], axis = 1)
        
        #set last column name to country name
        all_forecast_data.rename(columns={all_forecast_data.columns[-1]: forecast['country']}, inplace=True)

    #add dates to summary sheet
    for i in range(0,len(all_forecast_data)):
         #write index of all_forecast_table to date column
        for i in range(0,len(all_forecast_data)):
            summary['B'+str(3 + i + country_start_row)] = all_forecast_data.index[i].strftime("%Y-%m-%d")

            for j in range(0, all_forecast_data.shape[1]):
                summary[get_column_letter(j+3) + str(2 + country_start_row)] = forecast_results[j]['country'].split("_")[0]
                #format bold and fill background to light blue
                summary[get_column_letter(j+3) + str(2 + country_start_row)].fill = PatternFill("solid", fgColor="DCE6F1")
                #set alignment right
                summary[get_column_letter(j+3) + str(2 + country_start_row)].alignment = Alignment(horizontal='right')
                #format bold and fill background to light blue
                summary[get_column_letter(j+3) + str(2 + country_start_row)].font = Font(bold=True)
                summary[get_column_letter(j+3) + str(3 + i + country_start_row)] = round(all_forecast_data.iloc[i,j],2)
                #set width to fit content
                summary.column_dimensions[get_column_letter(j+3)].width = 12

    #add borders
    for row in summary.iter_rows(min_row=4, max_row=2 + country_start_row + len(all_forecast_data), min_col=2, max_col=2 + all_forecast_data.shape[1]):
                for cell in row:
                    cell.border = rrange_border
         

    for forecast in forecast_results:
       
        data = forecast['data']

        naive_rmse = forecast['naive_rmse']

        forecast_mean = forecast['forecast']

        

        #calculate bias corrected MAPE 
        if forecast['model'] == 'Random Forest w/ Lags':
            #check if bias corrected forecast is available
            forecast_corrected = None
            if forecast['bias_corrected'] is not None:
                forecast_corrected = forecast['bias_corrected']

                forecast_bias = forecast['bias_vec'].flatten()

                bias_corrected_residuals = forecast['ts_bias_corrected'] - forecast['test_actual']
                ts_bias_corrected = np.array(forecast['ts_bias_corrected'])
                # Avoid division by zero by masking out zero values
                nonzero_mask = test_actual != 0
                if np.any(nonzero_mask):
                    mape = 100 * np.mean(np.abs(ts_bias_corrected[nonzero_mask] - test_actual[nonzero_mask]) / np.abs(test_actual[nonzero_mask]))
                else:
                    mape = np.nan
                forecast['bias_corrected_mape'] = round(mape, 2)

            test_actual = np.array(forecast['test_actual'])
            
        country_start_row += 16

        #create a new worksheet for each country
        ws = wb.create_sheet(forecast['country'])
 
        #write forecast results to excel
        ws['A1'] = forecast['country'] + " - " + forecast['indicator'] + " - " + "Next 12 months forecast as of:  " + data.index[-1].strftime("%Y-%m-%d") 
        ws.merge_cells('A1:G1')
        ws['A3'] = "Model" 
        ws['B3'] = forecast['model'] +" - " + str(forecast['params'])
        ws['A4'] = "Training Data"
        ws['B4'] = data.index[0].strftime("%Y-%m-%d") + " to " + data.index[-12].strftime("%Y-%m-%d")
        ws['A5'] = "Test Data"
        ws['B5'] = data.index[-12].strftime("%Y-%m-%d") + " to " + data.index[-1].strftime("%Y-%m-%d")
        ws['A7'] = "Date"
        ws['B7'] = "Inflation Change YoY %"
        ws['C7'] = "Lower Bound - 95% Confidence Interval"
        ws['D7'] = "Upper Bound - 95% Confidence Interval"
        ws['E7'] = "Bias Corrected Forecast"





        #add exogenous variables in C3 
        if(forecast['predictors'] is not None):
              exogenous_variables = forecast['predictors'].tolist()
              exogenous_variables = ', '.join(exogenous_variables)	
              ws['C3'] = "Exogenous Variables"
              #formeat bold and fill background to light blue
              ws['C3'].font = Font(bold=True)
              ws['D3'] = exogenous_variables

              #add borders to exogenous variables
              for row in ws.iter_rows(min_row=3, max_row=3, min_col=3, max_col=4):
                for cell in row:
                    cell.border = rrange_border

        #format bold and fill background to light blue


        #format d4 and e4
        ws['C3'].font = Font(bold=True)
        ws['C3'].fill = PatternFill("solid", fgColor="DCE6F1")

      



        for i in range(0,len(forecast_mean)):
            ws['A'+str(8+i)] = forecast_mean.index[i].strftime("%Y-%m-%d")
            ws['B'+str(8+i)] = round(forecast_mean[i],2)
            ws['C'+str(8+i)] = round(forecast['conf_int']['lower'].iloc[i],2)
            ws['D'+str(8+i)] = round(forecast['conf_int']['upper'].iloc[i],2)
            if forecast['model'] == 'Random Forest w/ Lags':
                if (forecast_corrected is not None):
                    ws['E'+str(8+i)] = round(forecast_corrected.iloc[i],2)


        #add chart
        chart = Image(forecast['chart'])

        forecast_chart_list.append(forecast['chart'])


        ws.add_image(chart,'B'+str(len(forecast_mean)+10))


        #add metrics
        ws['A54'] = "Test Set"
        ws['A55'] = "RMSE" 
        ws['B55'] = str(forecast['test_metrics']['RMSE'])
        ws['A56'] = "Naive RMSE"  
        ws['B56'] = str(naive_rmse)
        ws['A57'] = "MAPE"
        ws['B57'] = str(forecast['test_metrics']['MAPE']) + "%"
        if forecast['model'] == 'Random Forest w/ Lags':
            if forecast_corrected is not None:
                ws['C54']= "Bias Corrected"
                ws['C55'] = str(forecast['rmse_ts'])
                ws['C56'] = str(naive_rmse)
                ws['C57']  = str(forecast['bias_corrected_mape']) + "%"


        #add test predictions
        ws['A59'] = "Test Predictions"
        ws['A60'] = "Date"
        ws['B60'] = "Predicted"
        ws['C60'] = "Actual"


        if forecast['model'] == 'Random Forest w/ Lags':
            ws['D60'] = "Bias Vector"
            ws['E60'] = 'Bias Corrected'

        #add test predictions for test set
        for i in range(len(forecast['test_predictions'])):
            ws['A'+str(61+i)] = forecast['test_predictions'].index[i].strftime("%Y-%m-%d")
            ws['B'+str(61+i)] = round(float(forecast['test_predictions'].iloc[i]), 2)
            ws['C'+str(61+i)] = round(float(forecast['test_actual'].iloc[i]), 2)

            if forecast['model'] == 'Random Forest w/ Lags':
                if (forecast_corrected is not None):
                    ws['D' + str(61+i)] = round(forecast_bias[i], 2)
                    ws['E' + str(61+i)] = round(forecast['ts_bias_corrected'][i], 2)


        #add test chart
        chart = Image(forecast['test_chart'])
        ws.add_image(chart,'B75')

        #if random forest, add importance chart 
        if forecast['model'] == 'Random Forest w/ Lags':
            chart = Image(forecast['importance_chart'])
            ws.add_image(chart,'F75')

       # Set the text in cells left-aligned
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column): 
            for cell in row: 
                cell.alignment = Alignment(horizontal='left')
    
        


        # set borders
        for row in ws.iter_rows(min_row=3, max_row=5, min_col=1, max_col=2):
                for cell in row:
                    cell.border = rrange_border

        for row in ws.iter_rows(min_row=7, max_row=19, min_col=1, max_col=5):
                for cell in row:
                        cell.border = rrange_border

        for row in ws.iter_rows(min_row=54, max_row=57, min_col=1, max_col=3):
                for cell in row:
                    cell.border = rrange_border

        for row in ws.iter_rows(min_row=59, max_row=72, min_col=1, max_col=5):
                for cell in row:
                    cell.border = rrange_border
        
        #set column width to fit content
        column_letters = tuple(get_column_letter(col_number + 1) for col_number in range(ws.max_column))
        for column_letter in column_letters:
            ws.column_dimensions[column_letter].width = 50

        #insert column 
        ws.insert_cols(1)
        ws.insert_rows(1)

        #set row width
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20

        #set worksheet background fill to white
        for row in ws.iter_rows(max_row = 300, max_col = 15):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FFFFFF")


        #set font and color for cell B2 to blue and bold
        ws['B2'].font = Font(color="366092", bold=True, size = 22)

        #set font and color for cells B4:B6 to bold
        for row in ws.iter_rows(min_row=4, max_row=6, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="DCE6F1")


        #set font and color for cells B8:E8 to bold
        for row in ws.iter_rows(min_row=8, max_row=8, min_col=2, max_col=6):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="DCE6F1")

        
        #set font and color for cells B49 to bold
        for row in ws.iter_rows(min_row=55, max_row=55, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(bold=True)

        #merge and center cells B4:B6
        ws.merge_cells('B55:C55')
        ws['B55'].alignment = Alignment(horizontal='center')
        ws['B55'].fill = PatternFill("solid", fgColor="DCE6F1")


        #merge and center cells B54:D55
        ws.merge_cells('B60:F60')
        ws['B60'].alignment = Alignment(horizontal='center')
        ws['B60'].fill = PatternFill("solid", fgColor="DCE6F1")
        ws['B60'].font = Font(bold=True)

        #merge and center cells B59:D59
        for row in ws.iter_rows(min_row=61, max_row=61, min_col=2, max_col=6):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="DCE6F1")

     
        #write model summary to excel
        if forecast['model'] not in ['VAR', 'CNN','Random Forest w/ Lags']:
            summary_table_0 = forecast['model_summary'].summary().tables[0].data
            summary_table_1 = forecast['model_summary'].summary().tables[1].data
            summary_table_2 = forecast['model_summary'].summary().tables[2].data

        ws.merge_cells('B102:C102')
        ws['B102'] = "Appendix A: Model Summary"
        ws['B102'].font = Font(bold=True)
        ws['B102'].fill = PatternFill("solid", fgColor="DCE6F1")

        #fill in model summary
        #general info
        ws['B103'] = "Model"
        if (forecast['model'] in ['VAR']):
            ws['C103'] = "VAR" + " - " + str(forecast['params'])
        elif (forecast['model'] in ['CNN', 'Random Forest w/ Lags']):
            ws['C103'] = forecast['model'] 
        else:
            ws['C103'] = summary_table_0[1][1]

        ws['B104'] = "Sample"
        if (forecast['model'] == 'VARMAX'):
            ws['C104'] = summary_table_0[5][1] + summary_table_0[6][1]
        elif (forecast['model'] in ['VAR', 'CNN', 'Random Forest w/ Lags']):
             ws['C104'] = data.index[0].strftime("%Y-%m-%d") + " to " + data.index[-1].strftime("%Y-%m-%d")
        else:
            ws['C104'] = summary_table_0[4][1] + summary_table_0[5][1]

        ws['B105'] = "No. of Observations"
        if (forecast['model'] in ['VAR', 'CNN', 'Random Forest w/ Lags']):
            ws['C105'] = str(len(data))
        else:
            ws['C105'] = str(summary_table_0[0][3])
        
        if (forecast['model'] == 'Random Forest w/ Lags'):
            ws['B106'] = 'Residuals'
        
        else:
            #Terms and P_values
            ws['B106'] = "Terms"
            ws['C106'] = "P-value (Statistical significance)"

        # format header
        ws['B106'].font = Font(bold=True)
        ws['C106'].font = Font(bold=True)

        ws['B106'].fill = PatternFill("solid", fgColor="DCE6F1")
        ws['C106'].fill = PatternFill("solid", fgColor="DCE6F1")

        row_counter = 106
        #format terms
        #get term significance from each term in varmax summary table
        if (forecast['model'] == 'VARMAX'):
             for i in range(1,len(summary_table_2)):
                ws['B' + str(106 + i)] = summary_table_2[i][0]
                ws['C' + str(106 + i)] = summary_table_2[i][4]

                if(float(summary_table_2[i][4]) < 0.05):
                    #format good
                    ws['C' + str(106 + i)].fill = PatternFill("solid", fgColor=fill_good)
                    ws['C' + str(106 + i)].font = Font(color=font_good)
                else:
                    #format bad
                    ws['C' + str(106 + i)].fill = PatternFill("solid", fgColor=fill_bad)
                    ws['C' + str(106 + i)].font = Font(color=font_bad)
        
        elif (forecast['model'] in ['VAR']):
            #leave p-values blank
             for i in range(1, len(forecast['model_summary'].pvalues.iloc[:,0])):
                ws['B' + str(106 + i)] = forecast['model_summary'].pvalues.index[i]
                ws['C' + str(106 + i)] = forecast['model_summary'].pvalues.iloc[i,0]

                if(float(forecast['model_summary'].pvalues.iloc[i,0]) < 0.05):
                    #format good
                    ws['C' + str(106 + i)].fill = PatternFill("solid", fgColor=fill_good)
                    ws['C' + str(106 + i)].font = Font(color=font_good)
                else:
                    #format bad
                    ws['C' + str(106 + i)].fill = PatternFill("solid", fgColor=fill_bad)
                    ws['C' + str(106 + i)].font = Font(color=font_bad)

                row_counter = row_counter + 1 
        elif (forecast['model'] in ['ARIMA', 'SARIMAX']):

            for i in range(1,len(summary_table_1)):
                ws['B' + str(106 + i)] = summary_table_1[i][0]
                ws['C' + str(106 + i)] = summary_table_1[i][4]

                if(float(summary_table_1[i][4]) < 0.05):
                    #format good
                    ws['C' + str(106 + i)].fill = PatternFill("solid", fgColor=fill_good)
                    ws['C' + str(106 + i)].font = Font(color=font_good)
                else:
                    #format bad
                    ws['C' + str(106 + i)].fill = PatternFill("solid", fgColor=fill_bad)
                    ws['C' + str(106 + i)].font = Font(color=font_bad)
        elif (forecast['model'] == 'Random Forest w/ Lags'):

            residuals = forecast['rf_residuals']
            ws['B' + str(107)] = "Mean"
            ws['B' + str(108)] = "St Dev"
            ws['B' + str(109)] = "Kurtosis"
            ws['B' + str(110)] = "Skewness"

            if forecast['bias_corrected'] is not None:

                ws['B' + str(112)] = "Bias-Corrected Residuals"	
                #make bold, blue background
                ws['B' + str(112)].font = Font(bold=True)
                ws['B' + str(112)].fill = PatternFill("solid", fgColor="DCE6F1")

                ws['B' + str(113)] = "Mean"
                ws['B' + str(114)] = "St Dev"
                ws['B' + str(115)] = "Kurtosis"
                ws['B' + str(116)] = "Skewness"

                ws['C' + str(113)] = np.mean(forecast['bi_co_residuals'])
                ws['C' + str(114)] = np.std(forecast['bi_co_residuals'])
                ws['C' + str(115)] = kurtosis(np.array(forecast['bi_co_residuals']),fisher=True, bias=False)
                ws['C' + str(116)] = skew(forecast['bi_co_residuals'], bias=False)






        
            ws['C' + str(107)] = np.mean(residuals)
            ws['C' + str(108)] = np.std(residuals.values)
            ws['C' + str(109)] = float(kurtosis(residuals))
            ws['C' + str(110)] = float(skew(residuals))



            row_counter = 110

        else:
            pass

        #add residuals chart to worksheet if random forest w/ lags
        if forecast['model'] == 'Random Forest w/ Lags':
            chart = Image(forecast['residuals_chart'])
            ws.add_image(chart,'F' + str(102))

        if forecast['model'] == 'VARMAX':
             row_counter = len(summary_table_2)
        elif forecast['model'] in ['VAR']:
            row_counter = len(forecast['model_summary'].pvalues)
        elif forecast['model'] in ['CNN', 'Random Forest w/ Lags']:
            row_counter = row_counter
        else:
            row_counter = len(summary_table_1)
        
        #add model assumptions header
        ws['B' + str(106 + row_counter)] = "Assumptions"
        ws['C' + str(106 + row_counter)] = "P-value"

        # format header
        ws['B' + str(106 + row_counter)].font = Font(bold=True)
        ws['C' + str(106 + row_counter)].font = Font(bold=True)

        ws['B' + str(106 + row_counter)].fill = PatternFill("solid", fgColor="DCE6F1")
        ws['C' + str(106 + row_counter)].fill = PatternFill("solid", fgColor="DCE6F1")
        
        # model assumptions
        if forecast['model'] == 'VARMAX':
            ws['B' + str(107 + row_counter)] = "Ljung-Box (L1)"
            ws['C' + str(107 + row_counter)] = summary_table_1[1][1]

            ws['B' + str(108 + row_counter)] = "Heteroskedasticity"
            ws['C' + str(108 + row_counter)] = summary_table_1[3][1]

            ws['B' + str(109 + row_counter)] = "Jarque-Bera"
            ws['C' + str(109 + row_counter)] = summary_table_1[1][3]

            ws['B' + str(110 + row_counter)] = "Skew"
            ws['C' + str(110 + row_counter)] = summary_table_1[2][3]

            ws['B' + str(111 + row_counter)] = "Kurtosis"
            ws['C' + str(111 + row_counter)] = summary_table_1[3][3]
        #elif forecast['model'] in ['VAR', 'CNN']:
        #     if forecast['params'] > 10:
        #         nlags = forecast['params'] + 1
        #     else:
        #         nlags = 10
        #     ws['B' + str(107 + row_counter)] = "Test Whiteness"
        #     ws['C' + str(107 + row_counter)] = forecast['model_summary'].test_whiteness(nlags = nlags).conclusion_str + ' - ' + str(forecast['model_summary'].test_whiteness(nlags = nlags).h0)

        #     ws['B' + str(108 + row_counter)] = "Test Normality"
        #     ws['C' + str(108 + row_counter)] = forecast['model_summary'].test_normality().conclusion_str + ' - ' + str(forecast['model_summary'].test_normality().h0)

        #     #format
        #     if float(forecast['model_summary'].test_normality().pvalue) <= 0.05:
        #         ws['C' + str(108 + row_counter)].fill = PatternFill("solid", fgColor=fill_bad)
        #         ws['C' + str(108 + row_counter)].font = Font(color=font_bad)
        #     else:
        #         ws['C' + str(108 + row_counter)].fill = PatternFill("solid", fgColor=fill_good)
        #         ws['C' + str(108 + row_counter)].font = Font(color=font_good)
           
        #    #format whiteness
        #     if float(forecast['model_summary'].test_whiteness(nlags = nlags).pvalue <= 0.05):
        #         ws['C' + str(107 + row_counter)].fill = PatternFill("solid", fgColor=fill_bad)
        #         ws['C' + str(107 + row_counter)].font = Font(color=font_bad)
        #     else:
        #         ws['C' + str(107 + row_counter)].fill = PatternFill("solid", fgColor=fill_good)
        #         ws['C' + str(107 + row_counter)].font = Font(color=font_good)
            
        elif forecast['model'] in ['SARIMAX, ARIMA']: 
            ws['B' + str(107 + row_counter)] = "Ljung-Box (L1)"
            ws['C' + str(107 + row_counter)] = summary_table_2[1][1]

            ws['B' + str(108 + row_counter)] = "Heteroskedasticity"
            ws['C' + str(108 + row_counter)] = summary_table_2[3][1]

            ws['B' + str(109 + row_counter)] = "Jarque-Bera"
            ws['C' + str(109 + row_counter)] = summary_table_2[1][3]

            ws['B' + str(110 + row_counter)] = "Skew"
            ws['C' + str(110 + row_counter)] = summary_table_2[2][3]

            ws['B' + str(111 + row_counter)] = "Kurtosis"
            ws['C' + str(111 + row_counter)] = summary_table_2[3][3]


        #interpret results
        if forecast['model'] in ['SARIMAX, ARIMA']:
            if (float(summary_table_2[1][1]) > 0.05):
                ws['D' + str(107 + row_counter)] = "Residuals are independent."
                #format good
                ws['C' + str(107 + row_counter)].fill = PatternFill("solid", fgColor=fill_good)
                ws['C' + str(107 + row_counter)].font = Font(color=font_good)
            else:
                ws['D' + str(107 + row_counter)] = "Residuals are not independent."
                #format bad
                ws['C' + str(107 + row_counter)].fill = PatternFill("solid", fgColor=fill_bad)
                ws['C' + str(107 + row_counter)].font = Font(color=font_bad)

        
            if (float(summary_table_2[3][1]) > 0.05):
                ws['D' + str(108 + row_counter)] = "Residuals have the same variance."
                #format good
                ws['C' + str(108 + row_counter)].fill = PatternFill("solid", fgColor=fill_good)
                ws['C' + str(108 + row_counter)].font = Font(color=font_good)
            else:
                ws['D' + str(108 + row_counter)] = "Residuals have different variances."
                #format bad
                ws['C' + str(108 + row_counter)].fill = PatternFill("solid", fgColor=fill_bad)
                ws['C' + str(108 + row_counter)].font = Font(color=font_bad)
                
            if (float(summary_table_2[1][3]) > 0.05):
                ws['D' + str(109 + row_counter)] = "Residuals are normally distributed."
                #format good
                ws['C' + str(109 + row_counter)].fill = PatternFill("solid", fgColor=fill_good)
                ws['C' + str(109 + row_counter)].font = Font(color=font_good)
            else:
                ws['D' + str(109 + row_counter)] = "Residuals are not normally distributed."
                #format bad
                ws['C' + str(109 + row_counter)].fill = PatternFill("solid", fgColor=fill_bad)
                ws['C' + str(109 + row_counter)].font = Font(color=font_bad)

            if (float(summary_table_2[2][3]) > 0):  
                ws['D' + str(110 + row_counter)] = "Positive skew - model underestimates the mean."
                #format neutral
                ws['C' + str(110 + row_counter)].fill = PatternFill("solid", fgColor=fill_neutral)
                ws['C' + str(110 + row_counter)].font = Font(color=font_neutral)
            elif (float(summary_table_2[2][3]) < 0):                                                                                                       
                ws['D' + str(110 + row_counter)] = "Negative skew - Model overestimates the mean."
            #format neutral
                ws['C' + str(110 + row_counter)].fill = PatternFill("solid", fgColor=fill_neutral)
                ws['C' + str(110 + row_counter)].font = Font(color=font_neutral)

            elif (np.abs(float(summary_table_2[2][3])) >= 2):
                ws['C' + str(110 + row_counter)] = "Significant skew - over/under-estimation."
                #format bad
                ws['C' + str(110 + row_counter)].fill = PatternFill("solid", fgColor=fill_bad)
                ws['C' + str(110 + row_counter)].font = Font(color=font_bad)

            if (np.abs(float(summary_table_2[3][3])) < 3):
                ws['D' + str(111 + row_counter)] = "Residuals do not contain more extreme outliers than in a normal distribution."
                #format good
                ws['C' + str(111 + row_counter)].fill = PatternFill("solid", fgColor=fill_good)
                ws['C' + str(111 + row_counter)].font = Font(color=font_good)
            else:
                ws['D' + str(111 + row_counter)] = "Residuals contain more extreme outliers than in a normal distribution."
                #format bad
                ws['C' + str(111 + row_counter)].fill = PatternFill("solid", fgColor=fill_bad)
                ws['C' + str(111 + row_counter)].font = Font(color=font_bad)

            #borders
        for row in ws.iter_rows(min_row=102, max_row=111 + row_counter, min_col=2, max_col=3):
                for cell in row:
                    cell.border = rrange_border


        #add plot of data with slice used highlighted 

        #add data chart
        data_chart = Image(forecast['orig_data_chart'])

        ws.add_image(data_chart,'B' + str(113 + row_counter))

        #add residuals summary statistics
        ws['B' + str(140 + row_counter + 1)] = "Appendix C: Residuals Analysis"
        ws['B' + str(140 + row_counter + 1)].font = Font(bold=True)
        ws['B' + str(140 + row_counter + 1)].fill = PatternFill("solid", fgColor="DCE6F1")

        ws['B' + str(140 + row_counter + 2)] = "Mean"
        if forecast['model'] in ['VARMAX','VAR']:
             ws['C' + str(140 + row_counter + 2)] = str(round(forecast['residuals'].iloc[:,0].mean(),2))
        elif forecast['model'] in ['CNN', 'Random Forest w/ Lags']:
            pass
        else: 
            ws['C' + str(140 + row_counter + 2)] = str(round(forecast['residuals'].mean(),2))
        


        #CALCULATE white noise
        from statsmodels.stats.diagnostic import acorr_ljungbox
        if forecast['model'] == 'VARMAX':
            ljung_box = acorr_ljungbox(forecast['residuals'].iloc[:,0], lags=12)
        elif forecast['model'] in ['VAR']:
            #test for white noise with durbin watson
            from statsmodels.stats.stattools import durbin_watson
            ws['B' + str(140 + row_counter + 3)] = "Errors AutoCorrelated?"
            #write each score in durbin watson to new row
            ws['C' + str(140 + row_counter + 3)] = str(durbin_watson(forecast['residuals'])[0])

            #interpret durbin watson
            if (durbin_watson(forecast['residuals'])[0] < 1.5):
                ws['D' + str(140 + row_counter + 3)] = "Positive Autocorrelation"
            elif (durbin_watson(forecast['residuals'])[0] > 2.5):
                ws['D' + str(140 + row_counter + 3)] = "Negative Autocorrelation"
            else:
                ws['D' + str(140 + row_counter + 3)] = "No Autocorrelation"

            #format durbin watson
            if (durbin_watson(forecast['residuals'])[0] < 1.5) or (durbin_watson(forecast['residuals'])[0] > 2.5):
                ws['D' + str(140 + row_counter + 3)].fill = PatternFill("solid", fgColor=fill_bad)
                ws['D' + str(140 + row_counter + 3)].font = Font(color=font_bad)
            else:
                ws['D' + str(140 + row_counter + 3)].fill = PatternFill("solid", fgColor=fill_good)
                ws['D' + str(140 + row_counter + 3)].font = Font(color=font_good)

        elif forecast['model'] in ['CNN', 'Random Forest w/ Lags']:
            pass
        else:
            ljung_box = acorr_ljungbox(forecast['residuals'], lags=12)

    
        if forecast['model'] == 'VARMAX':
            p_values = ljung_box['lb_pvalue']

            if (p_values < 0.05).any():
                ws['B' + str(140 + row_counter + 3)] = "White Noise?"
            else:
                ws['B' + str(140 + row_counter + 3)] = "White Noise?"
                ws['C' + str(140 + row_counter + 3)] = "Yes"


        
        #calculate normality and add results
        from scipy.stats import shapiro
        if forecast['model'] not in ['CNN', 'Random Forest w/ Lags']:
            shapiro_test = shapiro(forecast['residuals'])
            if (shapiro_test[1] < 0.05):
                ws['B' + str(140 + row_counter + 4)] = "Normal?"
                ws['C' + str(140 + row_counter + 4)] = "No"
            else:
                ws['B' + str(140 + row_counter + 4)] = "Normal?"
                ws['C' + str(140 + row_counter + 4)] = "Yes"

            #add border
            ws.merge_cells('B' + str(140 + row_counter + 1) + ':C' + str(140 + row_counter + 1))
            for row in ws.iter_rows(min_row=140 + row_counter + 1, max_row=140 + row_counter + 4, min_col=2, max_col=3):
                    for cell in row:
                        cell.border = rrange_border

        #color code white noise
        if forecast['model'] == 'VARMAX': 
            if (p_values < 0.05).any():
                ws['C' + str(140 + row_counter + 3)].fill = PatternFill("solid", fgColor=fill_bad)
                ws['C' + str(140 + row_counter + 3)].font = Font(color=font_bad)
            else:
                ws['C' + str(140 + row_counter + 3)].fill = PatternFill("solid", fgColor=fill_good)
                ws['C' + str(140 + row_counter + 3)].font = Font(color=font_good)

        
        #color code normality results
        if forecast['model'] not in ['CNN', 'Random Forest w/ Lags']:
            if (shapiro_test[1] < 0.05):
                ws['C' + str(140 + row_counter + 4)].fill = PatternFill("solid", fgColor=fill_bad)
                ws['C' + str(140 + row_counter + 4)].font = Font(color=font_bad)
            else:
                ws['C' + str(140 + row_counter + 4)].fill = PatternFill("solid", fgColor=fill_good)
                ws['C' + str(140 + row_counter + 4)].font = Font(color=font_good)

        
        #add plot diagnostics 

        #create folder to store diagnostics
        if not os.path.exists('diagnostics'):
            os.makedirs('diagnostics')

        diagnostics = forecast['plot_diagnostics'] or None

        #add title to diagnostics
        if diagnostics is not None:
            diagnostics.suptitle('Appendix D: Plot Diagnostics')
            diagnostics.tight_layout()
            diagnostics.set_size_inches(12, 6)

            #save diagnostics chart in diagnostics folder
            diagnostics.savefig(('diagnostics/diagnostics_' + forecast['country'] + '.png'))

            #add plot diagnostics chart image to excel
            diagnostics = Image(('diagnostics/diagnostics_' + forecast['country'] + '.png'))
            ws.add_image(diagnostics, 'B' + str(151 + row_counter))

            #add border to image
            img = PILImage.open(('diagnostics/diagnostics_' + forecast['country'] + '.png'))
            bordered_img = ImageOps.expand(img, border=1, fill='black')
            bordered_img.save(('diagnostics/diagnostics_' + forecast['country'] + '.png'))

     

        ''' 

        #plot acf of squared residuals - this is to check for ARCH
        from statsmodels.graphics.tsaplots import plot_acf
        acf_plot = plot_acf(forecast['residuals']**2, lags=12, alpha=0.05)
        acf_plot.figure.set_size_inches(12, 6)
        plt.suptitle('Appendix E: Autocorrelation Function of Squared Residuals')
        
        acf_plot.tight_layout()
        acf_plot.savefig(('acf_' + forecast['country'] + '.png'))

        #add border to image
        img2 = PILImage.open(('acf_' + forecast['country'] + '.png'))
        bordered_img2 = ImageOps.expand(img2, border=1, fill='black')
        bordered_img2.save(('acf_' + forecast['country'] + '.png'))

        #add acf chart image to excel
        acf = Image(('acf_' + forecast['country'] + '.png'))
        ws.add_image(acf, 'B' + str(184 + row_counter))


        '''

        #color code RMSE
        ws['C57'].fill = PatternFill("solid",fgColor=fill_neutral) #NEUTRAL
        ws['C57'].font = Font(color=font_neutral)

        if (float(forecast['test_metrics']['RMSE']) < naive_rmse):
            ws['C56'].fill = PatternFill("solid",fgColor=fill_good) #GOOD
            ws['C56'].font = Font(color=font_good)
        elif(float(forecast['test_metrics']['RMSE']) == naive_rmse):
            ws['C56'].fill = PatternFill("solid",fgColor=fill_neutral) #NEUTRAL
            ws['C56'].font = Font(color=font_neutral)
        else:
            ws['C56'].fill = PatternFill("solid",fgColor=fill_bad) #BAD
            ws['C56'].font = Font(color=font_bad)

        #Color code 
        if (float(forecast['test_metrics']['MAPE']) <= 20):
            ws['C58'].fill = PatternFill("solid",fgColor=fill_good) #GOOD
            ws['C58'].font = Font(color=font_good)
        else:
            ws['C58'].fill = PatternFill("solid",fgColor=fill_bad) #BAD
            ws['C58'].font = Font(color=font_bad)
 
    
        #color code tab color based on MASE
        if (float(forecast['test_metrics']['MAPE']) > 20):
            ws.sheet_properties.tabColor = fill_bad #BAD
        else:
            ws.sheet_properties.tabColor = fill_good #GOOD  

    #save the workbook if forecast results are not empty
    if len(forecast_results) > 0:
        FILE_PATH = r'reports' + f"\{today_year}\{today_month}\{today}\\"
        SAVE_PATH = f"{forecast_results[0]['indicator']}_forecast_report_{datetime.today().strftime('%Y-%m-%d_%H_%M')}.xlsx"

        #create directory if it doesn't exist
        if not os.path.exists(FILE_PATH):
            os.makedirs(FILE_PATH)
        
        #save file
        wb.save(FILE_PATH + SAVE_PATH)
        
        #open saved file
        subprocess.Popen([FILE_PATH + SAVE_PATH], shell=True)



    

